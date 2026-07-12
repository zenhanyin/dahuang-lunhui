from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def scalar(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def parse_bool(value):
    value = scalar(value)
    if value == "":
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "开", "开启"}


def parse_atom(value):
    text = str(value).strip()
    if text == "":
        return ""
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+\.\d+", text):
        return float(text)
    if text.lower() == "true":
        return True
    if text.lower() == "false":
        return False
    return text


def parse_object(value, field_name):
    value = scalar(value)
    if value == "":
        return None
    if isinstance(value, dict):
        return value

    text = str(value).strip()
    try:
        parsed = json.loads(text)
        if not isinstance(parsed, dict):
            raise ValueError
        return parsed
    except Exception:
        result = {}
        pairs = [part.strip() for part in re.split(r"[;\n，,]+", text) if part.strip()]
        for pair in pairs:
            if "=" not in pair:
                raise ValueError(f"{field_name} must use JSON or key=value pairs: {text}")
            key, raw = pair.split("=", 1)
            result[key.strip()] = parse_atom(raw)
        return result


def rows_by_header(sheet):
    header = [scalar(cell.value) for cell in sheet[1]]
    rows = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(scalar(value) == "" for value in row):
            continue
        item = {header[index]: scalar(value) for index, value in enumerate(row) if index < len(header)}
        item["_row"] = row_index
        rows.append(item)
    return rows


def build_config(workbook_path: Path):
    wb = load_workbook(workbook_path, data_only=True)

    required_sheets = {"Chapters", "Choices", "StatusLines"}
    missing = required_sheets.difference(wb.sheetnames)
    if missing:
        raise SystemExit(f"Missing sheets: {', '.join(sorted(missing))}")

    chapters = []
    choices_by_act = {}

    for row in rows_by_header(wb["Choices"]):
        act = int(row.get("Act", ""))
        choice = {
            "text": str(row.get("Text", "")),
            "to": str(row.get("To", "")),
        }
        requires = parse_object(row.get("Requires", ""), "Requires")
        effects = parse_object(row.get("Effects", ""), "Effects")
        if requires:
            choice["requires"] = requires
        if effects:
            choice["effects"] = effects
        if parse_bool(row.get("TalentToQi", "")):
            choice["talentToQi"] = True
        if scalar(row.get("SetPath", "")) != "":
            choice["setPath"] = str(row["SetPath"])
        if scalar(row.get("SetOrigin", "")) != "":
            choice["setOrigin"] = str(row["SetOrigin"])
        if scalar(row.get("Log", "")) != "":
            choice["log"] = str(row["Log"])
        if parse_bool(row.get("AdvanceAct", "")):
            choice["advanceAct"] = True

        if not choice["text"] or not choice["to"]:
            raise SystemExit(f"Choices row {row['_row']} needs Text and To")

        order = int(row.get("Order", 0) or 0)
        choices_by_act.setdefault(act, []).append((order, choice))

    for row in rows_by_header(wb["Chapters"]):
        act = int(row.get("Act", ""))
        chapter = {
            "name": str(row.get("Name", "")),
            "text": {"normal": str(row.get("TextNormal", ""))},
            "choices": [choice for _, choice in sorted(choices_by_act.get(act, []), key=lambda item: item[0])],
        }
        if not chapter["name"] or not chapter["text"]["normal"]:
            raise SystemExit(f"Chapters row {row['_row']} needs Name and TextNormal")
        chapters.append((act, chapter))

    status_lines = {}
    for row in rows_by_header(wb["StatusLines"]):
        key = str(row.get("Key", ""))
        text = str(row.get("Text", ""))
        if key and text:
            status_lines[key] = text

    art_pools = {}
    if "ArtPools" in wb.sheetnames:
        for row in rows_by_header(wb["ArtPools"]):
            scene = str(row.get("Scene", ""))
            image = str(row.get("Image", ""))
            if scene and image:
                art_pools.setdefault(scene, []).append(image)

    config = {
        "chapters": [chapter for _, chapter in sorted(chapters, key=lambda item: item[0])],
        "statusLines": status_lines,
    }
    if art_pools:
        config["artPools"] = art_pools
    return config


def main():
    parser = argparse.ArgumentParser(description="Convert story-config.xlsx to story-config.js")
    parser.add_argument("input", nargs="?", default=str(ROOT / "story-config.xlsx"))
    parser.add_argument("output", nargs="?", default=str(ROOT / "story-config.js"))
    args = parser.parse_args()

    config = build_config(Path(args.input))
    js = "window.DAHUANG_STORY_CONFIG = "
    js += json.dumps(config, ensure_ascii=False, indent=2)
    js += ";\n"

    output_path = Path(args.output)
    output_path.write_text(js, encoding="utf-8")
    print(f"Wrote {output_path} with {len(config['chapters'])} chapters.")


if __name__ == "__main__":
    main()
