from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / "story-config.xlsx"


CHAPTERS = [
    [0, "雷后入局", "雷劫散后，西陵旧议厅的水镜亮起。天机营、冰心堂、魍魉、云麓、荒火都派人入席，却没有一个人愿意先说真话。\n\n{{status}}\n\n你这一世不必急着选正邪，而要先决定从哪里撬开大荒的第一道缝。"],
    [1, "诸门相疑", "第一条线索已经落地，却把更多旧债牵了出来：铜门裂纹、城外疫民、北溟盐晶、沉船兵器、九天祭席，全都不像偶然。\n\n{{status}}\n\n此时的选择会决定谁愿意信你，也会决定谁开始怕你。"],
    [2, "北溟旧约", "你沿着线索抵达北溟边境。雪原下埋着幽都旧约，海风里带着东海神域的潮声；某些神明把人间叫作试炼，也把牺牲叫作秩序。\n\n{{status}}\n\n真相变得更大，也更危险。你可以公开、潜入、修行、救人，或把自己推到铜门前。"],
    [3, "东海神域", "东海神域并非天外净土。潮宫残碑上写着九天与幽都的旧盟，神使说轮回是补天之器，魍魉影主却说那只是好听的锁链。\n\n{{status}}\n\n这一幕会把你的修炼方式变成结局倾向：守护者、破局者、独行者，或被因果拖入幽都。"],
    [4, "天命收束", "本世所有线索收束在铜门之前：诸门新约、幽都旧契、东海神域、九天祭席，都等你给出最后答案。\n\n道德越高，飞升越稳；因果越重，越容易先入幽都重走人间。正邪不是罪名，选择如何使用力量才是。"],
]


CHOICES = [
    [0, 1, "逼问魍魉长老：上一世谁藏起裂纹", "shadow", "", "karma=4;fragments=1", "", "", "魍魉席后一阵沉默。有人承认，铜门裂纹不是第一次出现。", ""],
    [0, 2, "先救城外疫民，换冰心堂公开药账", "iceHeart", "", "virtue=12;years=-1", "", "", "药账里夹着一枚旧印，印文正与铜门边缘吻合。", ""],
    [0, 3, "让遗墨机关鸟拓下水镜裂纹", "jadeClue", "", "fragments=1;qi=10", "是", "", "机关鸟飞回时，爪上多了一粒东海盐晶。", ""],
    [0, 4, "借天机营兵符封锁铜门，逼各派留下证词", "greatCouncil", "karmaMax=35", "virtue=6;karma=6;fragments=1", "", "", "封锁不是仁慈，却让三个说谎的人同时露怯。", ""],
    [0, 5, "听云麓弟子推演星图，查裂纹对应的天象", "tribulation", "", "qi=18;years=-1", "是", "", "星图尽头不是九天，而是一片被抹去名字的海。", ""],
    [0, 6, "去荒火营查沉船兵器，不问来路先问用途", "copperGate", "", "qi=22;karma=6;years=-1", "", "", "沉船铁锈里藏着神域火纹，像是专为破门而铸。", ""],
    [0, 7, "不入议厅，先跟踪夜哭里的影步残象", "foxRoad", "", "memory=1;karma=3", "", "", "影步尽头留着半句誓词：救人者也会成为钥匙。", ""],

    [1, 1, "公开天书残页，迫诸门当场站队", "greatCouncil", "", "virtue=8;karma=3", "", "", "诸门终于不再谈风骨，只开始谈各自欠下的债。", ""],
    [1, 2, "夜入北溟，追玄晶留下的幽都旧约", "northSea", "", "fragments=1;years=-1", "", "", "北溟风雪里，你听见幽都并不承认自己是地狱。", ""],
    [1, 3, "独自贴近铜门，听清门内喊出的名字", "copperGate", "", "qi=16;karma=8", "是", "", "铜门没有打开，只把你的倒影换成了上一世的样子。", ""],
    [1, 4, "请冰心堂先验疫民魂灯，查谁动过生死簿", "iceHeart", "virtueMin=55", "virtue=14;fragments=1;years=-1", "", "", "魂灯没有灭，说明有人把活人提前写进了死册。", ""],
    [1, 5, "与弈剑旧友论剑三夜，换他带你进旧阵眼", "tribulation", "", "qi=24;years=-2", "是", "", "剑意洗去浮躁，也削开阵眼外层的伪装。", ""],
    [1, 6, "让魍魉暗网散出假消息，引蛇出洞", "shadow", "karmaMax=45", "karma=9;fragments=1", "", "", "这不是作恶，是用影子照出更深的影子。", ""],
    [1, 7, "护送一队凡人离开裂隙，暂缓追查", "death", "yearsMin=1;virtueMin=80", "virtue=18;memory=1;years=-1", "", "", "他们不知道真相，却会在下一世替你记得一个名字。", ""],

    [2, 1, "把北溟旧约带回议厅，要求诸门共证", "greatCouncil", "fragmentsMin=2", "virtue=10;karma=5", "", "", "旧约一出，最先拔剑的不是魔修，而是仙盟使者。", ""],
    [2, 2, "潜入幽都边市，查轮回簿被谁改写", "hell", "karmaMin=45", "memory=2;karma=-8", "", "", "幽都没有审你，只让你看见自己每次逃避的代价。", ""],
    [2, 3, "闭关炼化盐晶，冲击更高境界", "tribulation", "", "qi=38;years=-2", "是", "", "盐晶化作潮声，丹田里第一次出现神域门影。", ""],
    [2, 4, "救北溟雪灾里的敌派弟子，不问立场", "iceHeart", "", "virtue=20;karma=-6;years=-1", "", "", "对方醒来后只说一句：原来你不是为了赢。", ""],
    [2, 5, "借荒火战阵强行撬开铜门外环", "copperGate", "realmMin=2", "qi=30;karma=12;fragments=1;years=-2", "", "", "铜门震动，九天祭席上有一个无名位被点亮。", ""],
    [2, 6, "用魍魉影契换取神域使者真名", "heartDemon", "karmaMin=20", "fragments=1;karma=10;memory=1", "", "", "你得到了真名，也听见心魔学会了那人的声音。", ""],
    [2, 7, "请太虚观召灵问古，不让任何一派独占答案", "jadeClue", "virtueMin=60", "virtue=8;fragments=2;years=-1", "", "", "古灵只留下四字：补天有价。", ""],
    [2, 8, "若碎片已足，提前拼合旧约试探轮回边界", "breakWheel", "fragmentsMin=4;realmMin=3", "", "", "", "", ""],

    [3, 1, "登潮宫残碑，问东海神域为何沉默", "jadeClue", "", "fragments=1;qi=18;years=-1", "是", "", "残碑回答你的不是文字，而是一段被献祭的飞升记忆。", ""],
    [3, 2, "救下被当作祭品的神域侍童", "iceHeart", "virtueMin=70", "virtue=22;karma=-10", "", "", "侍童把潮宫暗门告诉你，却求你别把所有神明都当敌人。", ""],
    [3, 3, "假意接受神使册封，查九天祭席名单", "greatCouncil", "karmaMax=60", "karma=12;fragments=1", "", "", "你坐上了不该坐的位置，也看清了谁在数人间的命。", ""],
    [3, 4, "让心魔代你入梦，偷听祭席背后的交易", "heartDemon", "karmaMin=30", "memory=2;karma=14", "", "", "心魔带回真相，也带回一枚更像你的面具。", ""],
    [3, 5, "以修为硬闯九天门缝，先看飞升真假", "ascend", "realmMin=5", "", "", "", "", ""],
    [3, 6, "以诸门新约重订绝地天通", "breakWheel", "fragmentsMin=4;virtueMin=75", "", "", "", "", ""],
    [3, 7, "因果已重，先入幽都照见本心再谈飞升", "hell", "karmaMin=55", "", "", "", "", ""],
    [3, 8, "寿元将尽，把东海神域的坐标交给下一世", "death", "yearsMin=1", "memory=2;fragments=1;years=-99", "", "", "你没有失败，只是把终局推给了更强的自己。", ""],

    [4, 1, "以高德行登九天，公开祭席后的旧账", "ascend", "realmMin=5;virtueMin=90;karmaMax=45", "", "", "", "", ""],
    [4, 2, "以天命碎片重封铜门，结束献祭式轮回", "breakWheel", "fragmentsMin=5;virtueMin=70", "", "", "", "", ""],
    [4, 3, "牺牲本世修为，换诸门凡人全数撤离", "death", "virtueMin=100", "memory=3;fragments=1;qi=-80;years=-99", "", "", "你把胜利留给别人，把答案留给下一世。", ""],
    [4, 4, "带着因果强行飞升，赌九天不敢拒你", "ascend", "realmMin=6;karmaMin=45", "", "", "", "", ""],
    [4, 5, "若因果压过道心，入幽都重走人间", "hell", "karmaMin=65", "", "", "", "", ""],
    [4, 6, "以魍魉、冰心、天机三方证词重审旧约", "greatCouncil", "fragmentsMin=3;virtueMin=65", "virtue=12;karma=-8;memory=1", "", "", "这一次不是你一个人对抗天命，而是大荒一起开口。", ""],
    [4, 7, "什么都不争，只把真相刻进轮回记忆", "death", "", "memory=2;fragments=1;years=-99", "", "", "你选择下一世再来，但不是空手而来。", ""],
    [4, 8, "若境界已至巅峰，直面东海神域背后的神界边界", "ascend", "realmMin=6;fragmentsMin=4", "", "", "", "", ""],
]


STATUS_LINES = [
    ["highKarma", "你的因果太重，魍魉席后的影子始终盯着你的手；若再强行登天，幽都会先来要账。"],
    ["highVirtue", "你的名声已足以让诸门暂时放下私怨，但真正的信任仍要靠下一步换来。"],
    ["normal", "你尚未赢得所有人的信任，诸门只肯给你一次说明的机会。"],
]

ART_POOLS = [
    ["birth", 1, "assets/xiling-ruins.webp", "初始与西陵旧地"],
    ["birth", 2, "assets/great-council.webp", "议厅回忆"],
    ["memoryArchive", 1, "assets/copper-gate.webp", "记忆与铜门"],
    ["memoryArchive", 2, "assets/underworld-ledger.webp", "轮回档案"],
    ["sect", 1, "assets/xiling-ruins.webp", "门派修行"],
    ["iceHeart", 1, "assets/iceheart-hall.webp", "冰心堂"],
    ["iceHeart", 2, "assets/sect-new-pact.webp", "救人与联盟"],
    ["shadow", 1, "assets/leize-shadow.webp", "魍魉影线"],
    ["shadow", 2, "assets/cold-shadow.webp", "暗线追查"],
    ["foxRoad", 1, "assets/fox-road.webp", "夜路狐影"],
    ["jadeClue", 1, "assets/east-sea-tide-palace.webp", "神域线索"],
    ["jadeClue", 2, "assets/copper-gate.webp", "铜门裂纹"],
    ["copperGate", 1, "assets/copper-gate.webp", "铜门"],
    ["copperGate", 2, "assets/east-sea-godrealm.webp", "神域门影"],
    ["northSea", 1, "assets/beiming-pact.webp", "北溟旧约"],
    ["northSea", 2, "assets/beiming-road.webp", "北溟道路"],
    ["greatCouncil", 1, "assets/sect-new-pact.webp", "诸门新约"],
    ["greatCouncil", 2, "assets/great-council.webp", "旧议厅"],
    ["tribulation", 1, "assets/tribulation.webp", "雷劫破境"],
    ["afterBreak", 1, "assets/sect-new-pact.webp", "阶段推进"],
    ["afterBreak", 2, "assets/beiming-pact.webp", "北溟推进"],
    ["afterBreak", 3, "assets/east-sea-tide-palace.webp", "神域推进"],
    ["heartDemon", 1, "assets/heart-trial.webp", "心魔问道"],
    ["heartDemon", 2, "assets/underworld-ledger.webp", "因果照见"],
    ["hell", 1, "assets/underworld-ledger.webp", "幽都轮回簿"],
    ["hell", 2, "assets/hell-rebirth.webp", "幽都重生"],
    ["ascend", 1, "assets/true-ascend.webp", "飞升真相"],
    ["ascend", 2, "assets/east-sea-tide-palace.webp", "神界边界"],
    ["breakWheel", 1, "assets/copper-gate.webp", "重订天通"],
    ["breakWheel", 2, "assets/sect-new-pact.webp", "诸门共证"],
    ["trueEnding", 1, "assets/true-ascend.webp", "真结局"],
    ["sacrifice", 1, "assets/copper-gate.webp", "牺牲封门"],
    ["death", 1, "assets/xiling-ruins.webp", "重生回环"],
    ["death", 2, "assets/underworld-ledger.webp", "记忆沉淀"],
]


def clear_and_write(ws, rows):
    ws.delete_rows(1, ws.max_row)
    for row in rows:
        ws.append(row)


def style_sheet(ws, widths):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="26352A")
    header_font = Font(color="F7F1DE", bold=True)
    thin = Side(style="thin", color="CBD5C0")
    border = Border(bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def main():
    wb = load_workbook(BOOK)
    clear_and_write(wb["Chapters"], [["Act", "Name", "TextNormal"], *CHAPTERS])
    clear_and_write(wb["Choices"], [["Act", "Order", "Text", "To", "Requires", "Effects", "TalentToQi", "SetPath", "Log", "AdvanceAct"], *CHOICES])
    clear_and_write(wb["StatusLines"], [["Key", "Text"], *STATUS_LINES])
    if "ArtPools" not in wb.sheetnames:
        wb.create_sheet("ArtPools")
    clear_and_write(wb["ArtPools"], [["Scene", "Order", "Image", "Note"], *ART_POOLS])
    style_sheet(wb["Chapters"], [8, 18, 96])
    style_sheet(wb["Choices"], [8, 8, 46, 18, 30, 30, 12, 16, 62, 12])
    style_sheet(wb["StatusLines"], [20, 88])
    style_sheet(wb["ArtPools"], [22, 8, 38, 36])
    wb.save(BOOK)
    print(f"Expanded {BOOK} to {len(CHAPTERS)} chapters and {len(CHOICES)} choices.")


if __name__ == "__main__":
    main()
