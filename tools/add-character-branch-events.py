from pathlib import Path

from openpyxl import load_workbook


WORKBOOKS = [Path("story-config.xlsx"), Path("site-source/story-config.xlsx")]


EXTRA_EVENTS = [
    # Act 2: first night, people appear through immediate trouble.
    [2, 11, "替伏枫送一包冷针到药庐", "afterBreak", "", "iceLedger=2;virtue=6;karma=-1;years=-1", "", "", "伏枫把冷针交给你，只说若门后的人还在喘气，就别先问他站在哪边。", ""],
    [2, 12, "帮断不悔把盾车推上街口", "afterBreak", "", "virtue=8;wandererCred=1;qi=8;years=-1", "TRUE", "", "断不悔没有说谢，只把最轻的一面盾递给你，说别让孩子看见火。", ""],
    [2, 13, "在镜室外听见冷喻问话", "afterBreak", "", "shadowTrust=1;moFeiyunProof=1;karma=2", "", "", "冷喻隔着门问：如果影子不背叛人，人还会不会背叛影子？你没有立刻回答。", ""],
    [2, 14, "替风落收起一封没寄出的信", "afterBreak", "", "moFeiyunProof=1;virtue=3;karma=1", "", "", "风落看见你手里的信，神色微变。他让你别拆，只替他压在星灯下面。", ""],

    # Act 3: old objects begin pointing toward named relationships.
    [3, 9, "把机关鸟交还给遗墨", "afterBreak", "", "inkContract=3;fragments=1;virtue=4", "", "", "遗墨接过机关鸟，先检查它有没有受惊，再问你盐晶是从哪里滚出来的。", ""],
    [3, 10, "请玄素看一眼镜上的灰", "afterBreak", "", "shadowTrust=1;memory=1;karma=-1", "", "", "玄素没有擦镜，只在灰上添了一笔。镜中的影子忽然安静下来。", ""],
    [3, 11, "把药签拿给伏枫和清时同看", "afterBreak", "", "iceLedger=3;virtue=8;karma=-2", "", "", "伏枫看时辰，清时看药味。两人谁也没先说话，却同时把药炉挪近了些。", ""],
    [3, 12, "让断不悔辨认旧盾背面的手印", "afterBreak", "", "wandererCred=1;virtue=8;karma=-1", "", "", "断不悔用指节碰了碰手印，说这面盾不该被收进库里。", ""],

    # Act 4: roads branch through companions, not lore exposition.
    [4, 7, "陪玄晖在雪里停一停", "afterBreak", "", "xuanhuiBond=3;sunCinders=1;virtue=5;years=-1", "", "", "玄晖停在风里很久。他没有解释金灰，只问你冷不冷。", ""],
    [4, 8, "跟令狐九尾把商队带回避风处", "afterBreak", "", "foxBond=3;wandererCred=1;virtue=8;years=-1", "", "", "令狐九尾数错了三次人数，最后干脆把每个人的名字都念了一遍。", ""],
    [4, 9, "听青阳说一句黄泉到碧落", "afterBreak", "", "swordHeart=3;qingyangTrace=2;qi=8", "TRUE", "", "青阳说完那句旧话就闭了嘴。雪落在剑鞘上，没有化。", ""],
    [4, 10, "沿着莫非云留下的白羽走一段", "afterBreak", "", "moFeiyunProof=2;virtue=5;fragments=1;years=-1", "", "", "白羽落在冰面上，指向潮声更深的地方。莫非云没有回头，只抬手让你慢些。", ""],

    # Act 5: testimony scene lets characters speak, still partial.
    [5, 7, "让伏枫替清时补完药账", "afterBreak", "", "iceLedger=3;virtue=10;karma=-3", "", "", "伏枫提笔时，清时第一次没有抢话。药账上多出一个活人的名字。", ""],
    [5, 8, "让断不悔和令狐九尾先数人", "afterBreak", "", "wandererCred=2;foxBond=1;virtue=10;years=-1", "", "", "断不悔数军册，令狐九尾数商队。两边数字对不上，却救回了三个漏掉的人。", ""],
    [5, 9, "让冷喻和莫非云隔镜对话", "afterBreak", "", "moFeiyunProof=3;shadowTrust=2;karma=-2", "", "", "冷喻没有看水镜，只问莫非云是否后悔。莫非云摇头，像早就知道答案。", ""],
    [5, 10, "让夜哭把第二枚令牌翻过来", "afterBreak", "", "shadowTrust=3;karma=1;fragments=1", "", "", "夜哭把令牌翻过来，背面没有血，只有一道被旧刀磨平的刻痕。", ""],
    [5, 11, "让遗墨拆开机关鸟腹中的盐晶", "afterBreak", "", "inkContract=3;fragments=1;virtue=5", "", "", "遗墨拆得很慢，像怕弄疼那只鸟。盐晶里落出一粒细小的铜屑。", ""],
]


def update(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb["Choices"]
    headers = [cell.value for cell in ws[1]]
    idx = {name: headers.index(name) + 1 for name in headers if name}

    existing = {
        (ws.cell(row, idx["Act"]).value, ws.cell(row, idx["Order"]).value)
        for row in range(2, ws.max_row + 1)
    }
    for row in EXTRA_EVENTS:
        key = (row[0], row[1])
        if key not in existing:
            ws.append(row)

    wb.save(path)
    print(f"added character branch events to {path}")


for workbook in WORKBOOKS:
    update(workbook)
