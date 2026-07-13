from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOKS = [ROOT / "story-config.xlsx"]


CHAPTERS = [
    (
        0,
        "这一世从何处醒来",
        "你睁眼时，掌心压着一小片温热的铜锈。\n\n"
        "窗纸外透着灰白的光。墙根下传来压低的哭声，雾里有一点灯火晃动；床边的药炉还温着，炉灰被人拨开过。\n\n"
        "{{status}}\n\n"
        "你还不知道自己为什么会回来。先活下去。",
    ),
    (
        1,
        "门前",
        "十二岁到十六岁之间，你走到许多门前。\n\n"
        "门房把木牌推到你面前，药庐的童子在阶下等你回话。桌角还压着一张空白路引，墨没有干。\n\n"
        "{{status}}\n\n"
        "你不必马上相信谁。先决定今晚住在哪里。",
    ),
    (
        2,
        "雨夜有人敲门",
        "入夜后，雨把街上的脚印冲淡。\n\n"
        "你听见三处声音：药庐后门有人咳血，城墙外有小孩哭，巷口的灯自己灭了又亮。\n\n"
        "{{status}}\n\n"
        "门没有锁。你可以出去，也可以假装没有听见。",
    ),
    (
        3,
        "旧物开口",
        "你带回来的东西被放在桌上。\n\n"
        "半片铜锈、湿透的羽片、无字药签、断掉的剑穗，每一样都很轻，却让屋里的人不再说话。\n\n"
        "{{status}}\n\n"
        "有人伸手想拿走它们。你按住了桌沿。",
    ),
    (
        4,
        "雪线之外",
        "北边的雪落得很慢。\n\n"
        "雪地里有一串脚印，旁边落着不融的金灰。远处有门，门前有灯。\n\n"
        "{{status}}\n\n"
        "你还不知道门后是谁，只知道有人已经等得太久。",
    ),
    (
        5,
        "水镜前",
        "水镜亮起时，屋里没有人先说话。\n\n"
        "药账、羽片、断剑、盐晶和旧令依次摆开。每个人都看见了自己不想看的那一部分。\n\n"
        "{{status}}\n\n"
        "这一次，轮到他们回答你。",
    ),
    (
        6,
        "门开之前",
        "铜门前没有风。\n\n"
        "你听见身后有人拔剑，有人点灯，有人压低声音叫你的名字。门上的铜锈一点点发热。\n\n"
        "你仍然不知道门后会有什么。可你知道，退回去的人也会被这扇门追上。",
    ),
]


BIRTH_CHOICES = [
    ("在西陵废墙下醒来", "西陵废墙", "virtue=8;memory=1;years=-1", "孩子的手印留在旧盾背面。你把盾翻过来，先听见墙外有人在哭。"),
    ("在江南药庐醒来", "江南药庐", "virtue=6;iceLedger=1;years=-1", "药炉快熄了。床边放着一张无字药签，纸角还带着血腥味。"),
    ("在雷泽沉船边醒来", "雷泽沉船", "karma=3;shadowTrust=1;memory=1", "雾里有人把半枚旧令丢进水里。令牌没有沉，只映出你的脸。"),
    ("在望星台醒来", "望星台", "fragments=1;qi=12;moFeiyunProof=1;years=-1", "星图尽头不是天，是一片被潮声擦过的黑。你捡到一根白羽。"),
    ("在旧雨亭醒来", "旧雨亭", "swordHeart=1;qingyangTrace=1;qi=10", "雨水从亭檐往上流。柱子上刻着一句旧话，你只认得其中半句。"),
    ("在白羽林醒来", "白羽林", "inkContract=1;fragments=1;qi=8", "一只破损的机关鸟倒在树根边。鸟腹里滚出一粒盐晶。"),
    ("在焦土旧营醒来", "焦土旧营", "fireOath=1;qi=18;karma=5;years=-1", "试锋石裂开三寸。地底传来笑声，你把耳朵移开。"),
    ("在青丘野渡醒来", "青丘野渡", "foxBond=1;virtue=7;memory=1", "渡口没有船。一个书生把最后半块干粮递给孩子，又匆匆走进雾里。"),
    ("在北溟雪线醒来", "北溟雪线", "xuanhuiBond=1;sunCinders=1;virtue=4;years=-1", "追魂符落在雪上没有燃。远处有人回头看了你一眼，又把脸藏进风里。"),
    ("在幽都井口醒来", "幽都井口", "memory=2;karma=-5;years=-1", "井口没有锁链，只有一盏灯被推到你面前。灯里照出你没救走的人。"),
    ("在白水渡口醒来", "白水渡口", "virtue=5;wandererCred=1;memory=1", "渡船还没靠岸，船夫却已经知道你要过河。他问你带不带旧账。"),
    ("在赤水河滩醒来", "赤水河滩", "qi=10;karma=2;fragments=1", "河沙里埋着半片铜符，摸上去像晒过很久。水下有人敲了三下石头。"),
    ("在鹿鸣古道醒来", "鹿鸣古道", "luck=1;virtue=4;years=-1", "古道两侧没有行人，只有鹿蹄印一路通向山外。你忽然不急着进任何门。"),
    ("在东海盐滩醒来", "东海盐滩", "fragments=1;qingyangTrace=1;karma=1", "潮水退去后，盐壳下露出一行旧字。海风磨平了大半，只剩一个字还清楚。"),
]


CHOICES_BY_ACT = {
    1: [
        ("去天机营报到", "greatCouncil", "", "virtue=6;fragments=1;years=-1", "天机营的人先给你一碗冷水，再给你一卷没有写完的军册。", "天机营"),
        ("去冰心堂做药童", "iceHeart", "", "iceLedger=2;virtue=8;karma=1", "药童要先学会熬夜。第一夜，你记住了三个人的咳声。", "冰心堂"),
        ("去弈剑听雨阁试剑", "tribulation", "", "swordHeart=2;qingyangTrace=1;qi=18;years=-1", "剑阁没有问你的来历，只让你在雨里站到天亮。", "弈剑听雨阁"),
        ("去云麓仙居守星灯", "jadeClue", "", "moFeiyunProof=1;virtue=5;fragments=1", "星灯很轻，灯芯却像一直在看你。", "云麓仙居"),
        ("去太虚观扫镜室", "heartDemon", "", "shadowTrust=1;memory=1;karma=-1", "镜室里的影子比你慢半步。你没有立刻叫人。", "太虚观"),
        ("去魍魉旧巷领夜牌", "shadow", "", "shadowTrust=2;karma=4;fragments=1", "夜牌冰凉。递牌的人提醒你，第一晚最好别问名字。", "魍魉"),
        ("去翎羽山庄修机关", "jadeClue", "", "inkContract=2;fragments=1;qi=8;years=-1", "机关房里堆满坏掉的鸟。你先修最小的一只。", "翎羽山庄"),
        ("去荒火营地搬炭", "copperGate", "", "fireOath=2;virtue=4;qi=18;karma=2", "炭火烫手。教头看见你没丢下木筐，才点了点头。", "荒火营地"),
        ("不入门，随商队走", "foxRoad", "", "wandererCred=2;virtue=6;memory=1", "商队缺一个会记账的人。你说自己可以试试。", "游侠"),
        ("留在城里做杂役", "greatCouncil", "", "wandererCred=3;virtue=5;karma=-2;memory=1", "杂役听得见很多门里听不见的事。你把扫帚靠在门边。", "游侠"),
    ],
    2: [
        ("去药庐后门", "iceHeart", "", "iceLedger=1;virtue=12;karma=-3;years=-1", "后门的灯快灭了。你扶住一个咳血的人，没有问他犯过什么错。", None),
        ("去城墙下", "foxRoad", "", "foxBond=2;virtue=9;years=-1", "墙下的孩子抱着一只湿透的狐尾铃。你先把外衣披给他。", None),
        ("去巷口看灯", "shadow", "", "shadowTrust=2;virtue=4;karma=2;fragments=1", "灯灭时，墙上多出第二个人影。你没有拔剑。", None),
        ("回屋守住桌上的旧物", "greatCouncil", "", "virtue=8;fragments=1", "有人从窗外伸手。你按住桌沿，听见对方轻轻叹气。", None),
        ("跟着机关鸟上屋檐", "jadeClue", "", "inkContract=2;virtue=8;fragments=1;years=-1", "机关鸟飞得歪歪斜斜，却总能避开巡夜人的灯。", None),
        ("把铜锈藏进袖中", "copperGate", "", "qi=14;karma=4", "铜锈贴着皮肤发热。你忍住没有把它交出去。", None),
        ("去雨亭听剑声", "tribulation", "", "swordHeart=2;qingyangTrace=1;qi=12", "亭中无人，剑声却从脚下传来。", None),
        ("沿雪味往北走", "northSea", "", "xuanhuiBond=1;sunCinders=1;years=-1", "雨里夹着雪味。你走了很久，才看见前方有人点灯。", None),
        ("先叫醒邻屋的人", "greatCouncil", "", "virtue=10;wandererCred=1;karma=-2", "有人骂你多管闲事。可火烧起来时，骂声先停了。", None),
        ("装作睡着", "heartDemon", "", "memory=1;karma=6", "脚步声停在床前很久。你闭着眼，听见铜锈在掌心跳了一下。", None),
    ],
    3: [
        ("按住半片铜锈", "copperGate", "", "fragments=1;qi=12;karma=2", "铜锈没有动，桌上的烛火却一起偏向门外。", None),
        ("递出无字药签", "iceHeart", "", "iceLedger=2;virtue=8", "药签被水一浸，背面慢慢浮出三个时辰。", None),
        ("摊开湿透的羽片", "jadeClue", "", "moFeiyunProof=2;virtue=6;fragments=1", "羽片上的盐痕没有干，像刚从海风里取下。", None),
        ("把断剑穗挂回墙上", "tribulation", "", "swordHeart=2;qingyangTrace=2;qi=10", "剑穗自己转了半圈，指向北边。", None),
        ("把旧令扣在灯下", "shadow", "", "shadowTrust=2;karma=3", "令牌投下的影子不止一枚。有人在暗处换了呼吸。", None),
        ("请所有人先坐下", "greatCouncil", "", "virtue=10;wandererCred=1;karma=-2", "屋里终于安静下来。你第一次听清雨声之外的心跳。", None),
        ("让机关鸟自己选择", "jadeClue", "", "inkContract=2;fragments=1", "机关鸟啄了啄盐晶，又啄了啄你的袖口。", None),
        ("把旧物全收起来", "heartDemon", "", "memory=1;karma=6;qi=8", "没有人敢拦你。可走出门时，你背后多了一道影子。", None),
    ],
    4: [
        ("走向雪地里的灯", "northSea", "", "xuanhuiBond=2;sunCinders=1;virtue=4;years=-1", "灯旁的人没有回头，只把一件斗篷丢给你。"),
        ("摸一摸不融的金灰", "northSea", "", "sunCinders=2;qi=16;karma=2", "金灰不烫，却让你想起很远很远的日光。"),
        ("绕开那扇门", "foxRoad", "", "wandererCred=1;virtue=8;years=-1", "你没有逞强。雪线外还有活人等着回家。"),
        ("把断剑穗埋进雪里", "tribulation", "", "swordHeart=2;qingyangTrace=1;karma=-2", "雪下传来一声轻响，像有人终于把剑放下。"),
        ("沿冰下潮声前行", "jadeClue", "", "fragments=1;moFeiyunProof=1;years=-1", "潮声从冰下传来。它不像邀请，更像提醒。"),
        ("回身护住后来的人", "greatCouncil", "", "virtue=12;karma=-4;wandererCred=1", "后来的人走得很慢。你把风挡在前面。"),
    ],
    5: [
        ("先让药庐的人开口", "iceHeart", "", "iceLedger=2;virtue=10;karma=-3", "药庐的人低头很久，终于报出一个没写进账里的名字。"),
        ("先让执令的人开口", "shadow", "", "shadowTrust=2;karma=2;fragments=1", "执令的人没有辩解，只把袖中的第二枚令牌放上桌。"),
        ("先让观星的人开口", "jadeClue", "", "moFeiyunProof=2;inkContract=1;fragments=1", "星图被推到水镜前，缺口正好对着东边。"),
        ("先让执剑的人开口", "tribulation", "", "swordHeart=2;qingyangTrace=2;qi=12", "执剑的人说自己只路过人间。可他的手一直按着剑鞘。"),
        ("不审谁，先救城外的人", "foxRoad", "", "virtue=14;wandererCred=2;years=-1", "水镜前少了你的声音，城外却多了几盏灯。"),
        ("把所有旧物带走", "heartDemon", "", "memory=2;karma=8;qi=12", "你离开时，没有人追上来。水镜里的影子却跟着你走。"),
    ],
    6: [
        ("推门前先回头看一眼", "trueAscend", "virtueMin=70", "virtue=8;memory=1", "你回头时，看见有人终于敢把灯举高。"),
        ("把铜锈贴在门缝上", "breakWheel", "fragmentsMin=4", "fragments=1;memory=2", "铜锈贴上门缝，里面传来一声很轻的叹息。"),
        ("用剑鞘抵住门", "sacrifice", "swordHeartMin=4", "virtue=6;memory=1", "剑没有出鞘。门却停了一息。"),
        ("独自走进门里", "falseAscend", "qiMin=160", "memory=2;karma=8", "门后的光太亮，你听不清身后的人在喊什么。"),
        ("退后，先把人带走", "death", "", "virtue=10;memory=1;years=-99", "你没有打开门。至少这一夜，城里还有人活着。"),
    ],
}


STATUS_LINES = [
    ("normal", "你还没有确定谁可信。"),
    ("highVirtue", "有人愿意把背后交给你，但你知道这份信任很薄。"),
    ("highKarma", "你手上沾过太多旧事，连灯火都离你远了一点。"),
]


ART_POOLS = [
    ("birth", 1, "assets/xiling-ruins.webp", "废墙、城外、旧梦"),
    ("birth", 2, "assets/fox-road.webp", "渡口、古道、野外"),
    ("birth", 3, "assets/leize-shadow.webp", "雾、水、沉船"),
    ("afterBreak", 1, "assets/great-council.webp", "门前与室内压迫感"),
    ("afterBreak", 2, "assets/iceheart-hall.webp", "药庐与救人事件"),
    ("afterBreak", 3, "assets/leize-shadow.webp", "夜路、巷口、旧令"),
    ("afterBreak", 4, "assets/fox-road.webp", "城外与游侠路径"),
    ("afterBreak", 5, "assets/beiming-road.webp", "雪线之外"),
    ("afterBreak", 6, "assets/copper-gate.webp", "旧物、铜锈、门缝"),
    ("afterBreak", 7, "assets/east-sea-godrealm.webp", "潮声、盐晶、东海边界"),
    ("trueAscend", 1, "assets/true-ascend.webp", "回头之后的天光"),
    ("trueAscend", 2, "assets/great-council.webp", "人间仍在身后"),
    ("breakWheel", 1, "assets/copper-gate.webp", "铜锈发热"),
    ("breakWheel", 2, "assets/east-sea-godrealm.webp", "门后的潮声"),
    ("sacrifice", 1, "assets/copper-gate.webp", "门停一息"),
    ("sacrifice", 2, "assets/tribulation.webp", "强撑的代价"),
    ("falseAscend", 1, "assets/copper-gate.webp", "白光吞声"),
    ("falseAscend", 2, "assets/heart-trial.webp", "冷光与失重"),
    ("death", 1, "assets/xiling-ruins.webp", "一世将尽"),
    ("death", 2, "assets/beiming-road.webp", "最后的风"),
    ("death", 3, "assets/hell-rebirth.webp", "再醒之前"),
    ("hell", 1, "assets/hell-rebirth.webp", "井口无声"),
    ("hell", 2, "assets/cold-shadow.webp", "水中倒影"),
]


def reset_sheet(ws, rows):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row in rows:
        ws.append(row)


def rewrite(path: Path) -> None:
    wb = load_workbook(path)

    reset_sheet(wb["Chapters"], CHAPTERS)

    choice_rows = []
    for order, (text, origin, effects, log) in enumerate(BIRTH_CHOICES, start=1):
        choice_rows.append([0, order, text, "afterBreak", "", effects, "TRUE" if "qi=" in effects else "", "未定", log, "", origin])

    for act, rows in CHOICES_BY_ACT.items():
        for order, item in enumerate(rows, start=1):
            text, to, requires, effects, log = item[:5]
            if act < 6:
                to = "afterBreak"
            set_path = item[5] if len(item) > 5 and item[5] else ""
            choice_rows.append([act, order, text, to, requires, effects, "TRUE" if "qi=" in effects else "", set_path, log, "", ""])

    reset_sheet(wb["Choices"], choice_rows)
    reset_sheet(wb["StatusLines"], STATUS_LINES)
    if "ArtPools" in wb.sheetnames:
        reset_sheet(wb["ArtPools"], ART_POOLS)

    wb.save(path)
    print(f"rewrote {path}")


for workbook in WORKBOOKS:
    rewrite(workbook)
