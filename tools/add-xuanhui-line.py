from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / "story-config.xlsx"

XUANHUI_CHOICES = [
    [0, 8, "追上自称玄晖的旅人，问他为何独往北溟", "northSea", "", "xuanhuiBond=1;sunCinders=1;fragments=1;years=-1", "", "", "玄晖不答，只说北溟有一个人欠他答案。水镜里一瞬映出十轮残日。", ""],
    [0, 9, "替玄晖挡下天机营盘问，只记下他的金色瞳纹", "foxRoad", "karmaMax=35", "xuanhuiBond=2;virtue=4;sunCinders=1", "", "", "他第一次认真看你，像在确认你是否把失势神子也当作活人。", ""],
    [1, 8, "查玄晖留在药账里的乌羽灰", "iceHeart", "", "xuanhuiBond=1;sunCinders=2;virtue=6", "", "", "乌羽灰遇药灯不燃，反而照出灵巫宫旧歌里的东君日轮。", ""],
    [1, 9, "把玄晖的旧名交给诸门，换取东皇太一档案", "greatCouncil", "karmaMax=50", "sunCinders=2;karma=8;xuanhuiBond=-1", "", "", "诸门只听见东皇太一四字，没人问玄晖为何宁愿用人间化名。", ""],
    [2, 9, "在北溟雪夜听玄晖讲射日之后", "northSea", "xuanhuiBondMin=2", "xuanhuiBond=2;sunCinders=2;memory=1;years=-1", "", "", "他说十日坠落时，他才知道贵为神明也会被命运写定。", ""],
    [2, 10, "帮玄晖摆脱化生魔禁制，换他借金乌火炼契", "tribulation", "sunCindersMin=3", "qi=32;karma=6;sunCinders=1;xuanhuiBond=1", "是", "", "金乌火没有焚尽幽都，只烧掉了旧契上最不肯见人的一角。", ""],
    [3, 9, "陪玄晖去见幽都王颛顼，哪怕再次被拒", "hell", "xuanhuiBondMin=4", "xuanhuiBond=2;sunCinders=1;memory=1;karma=-4", "", "", "颛顼没有开门。玄晖站在门外很久，仍没有低头。", ""],
    [3, 10, "释放金乌烬，让十日残光照穿神域祭席", "heartDemon", "sunCindersMin=5", "fragments=1;sunCinders=2;karma=10", "", "", "十道残光照出九天祭席后的空名，你也听见心魔学会了神的语气。", ""],
    [4, 9, "以玄晖之名重写东皇旧契：神子也有选择命运的权利", "breakWheel", "xuanhuiBondMin=6;sunCindersMin=6;virtueMin=70", "virtue=18;fragments=1;memory=2", "", "", "玄晖没有登回神座。他把最后一缕金乌火交给人间诸门共守。", ""],
    [4, 10, "强迫玄晖归位东皇，以十日之火焚开天门", "ascend", "sunCindersMin=6;karmaMin=35", "qi=60;karma=24;xuanhuiBond=-3", "是", "", "十日同升，天门大开。可你分不清那是飞升，还是又一次把他献上祭席。", ""],
]

LORE_ROWS = [
    ["玄晖", "文档依据", "大荒人物志：东皇太一失去力量后，以人类身份行动，取化名玄晖；玄为天，晖为日，玄晖即太阳别称。"],
    ["东皇太一", "身世", "大荒人物志：东皇太一是帝江之子、十日之首；帝俊对外宣称太阳十子为自己和羲和之子，并遮蔽真实身世。"],
    ["十日旧劫", "射日之后", "大荒人物志：太阳十子得知帝俊与帝江真相后化为金乌炙烤大地，最终被射落；太一在坠落时意识到自己的命运早被安排。"],
    ["北溟寻父", "人物动机", "大荒人物志：太一想找到生父帝江，也就是如今的幽都王颛顼；但颛顼拒绝回答，也拒绝提供面见机会。"],
    ["玄晖", "人物线定位", "本游戏把玄晖线写成人性与神格的冲突：尊重玄晖作为人的选择会提升玄晖缘，逼迫他归位东皇会提升力量但增加因果。"],
    ["金乌十日", "母题线", "天下，我们的大荒：帝俊、帝江、东皇太一皆可视为东夷太阳图腾脉络；游戏以金乌烬表现十日旧劫残留。"],
    ["数值", "xuanhuiBond", "玄晖缘。代表玩家理解、保护、尊重玄晖作为人的程度。高值开放新约/真结局辅助线。"],
    ["数值", "sunCinders", "金乌烬。代表玩家触及东皇太一与十日旧劫真相的深度。高值开放强力但高风险选项。"],
]


def rows(ws):
    return list(ws.iter_rows(min_row=2, values_only=True))


def style_sheet(ws, widths):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="26352A")
    header_font = Font(color="F7F1DE", bold=True)
    thin = Side(style="thin", color="CBD5C0")
    border = Border(bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def main():
    wb = load_workbook(BOOK)
    choices = wb["Choices"]
    header = [cell.value for cell in choices[1]]
    kept = [row for row in rows(choices) if not (row[2] and "玄晖" in str(row[2]))]
    choices.delete_rows(2, choices.max_row)
    for row in sorted([*kept, *XUANHUI_CHOICES], key=lambda r: (int(r[0]), int(r[1]))):
        choices.append(row)
    style_sheet(choices, [8, 8, 48, 18, 34, 36, 12, 16, 66, 12])

    if "CharacterLore" not in wb.sheetnames:
        wb.create_sheet("CharacterLore")
    lore = wb["CharacterLore"]
    lore.delete_rows(1, lore.max_row)
    lore.append(["Character", "Topic", "Notes"])
    for row in LORE_ROWS:
        lore.append(row)
    style_sheet(lore, [18, 22, 100])

    if "README" in wb.sheetnames:
        readme = wb["README"]
        existing = {row[0].value for row in readme.iter_rows(min_row=2, max_col=1)}
        if "人物线" not in existing:
            readme.append(["人物线", "CharacterLore 记录人物设定；Choices 可使用 xuanhuiBond / sunCinders 等数值做 Effects 和 Requires。"])

    wb.save(BOOK)
    print(f"Added Xuanhui line: {len(XUANHUI_CHOICES)} choices.")


if __name__ == "__main__":
    main()
